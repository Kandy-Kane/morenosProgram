from typing import List
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.cell import cell
from openpyxl.cell.cell import Cell, ERROR_CODES
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, alignment
from openpyxl.styles.fills import fills
from openpyxl.worksheet.dimensions import SheetDimension
from openpyxl.utils import get_column_letter
import time
import os.path
from os import error, path
# from tkinter import messagebox
import traceback
import sys
import datetime
from PyQt5 import QtCore, QtGui,QtWidgets,uic
from PyQt5.QtWidgets import QMainWindow,QApplication, QWidget
from PyQt5.QtCore import QObject, QThread, pyqtSignal,pyqtSlot
import logging
import threading
import time
import globalVars


thick_border_blue = Border(left=Side(style='thick',color='0066CC'), 
                    right=Side(style='thick',color='0066CC')   
                ) 

thick_border_blue_topBottom = Border(top=Side(style='thick',color='0066CC'), 
                    bottom=Side(style='thick',color='0066CC')   
                )                    

thin_border_all = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin'))

thin_border_all_grey = Border(left=Side(style='thin',color="DDDDDD"), 
                    right=Side(style='thin',color="DDDDDD"), 
                    top=Side(style='thin',color="DDDDDD"), 
                    bottom=Side(style='thin',color="DDDDDD"))

thin_border_sides = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style=None), 
                    bottom=Side(style=None))


thin_border_sides_Bottom = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style=None), 
                    bottom=Side(style='thin'))
ft1 = Font(name='Arial',bold=True, size=14)

align = Alignment(horizontal='center')
thick_border = Border(left=Side(style='thick'), 
                    right=Side(style='thick'), 
                    top=Side(style='thick'), 
                    bottom=Side(style='thick'))
# import main


    #GLOBAL ERROR VARIABLES
fileerrorLabel = None
qualerrorLabel = None
blockerrorLabel = None
starterrorLabel = None
enderrorLabel = None
monthCheck = False


class Worker4(QObject):
    finished = pyqtSignal()
    progress = pyqtSignal(int)
    found = pyqtSignal()


    def signOutSearch(self):
        try:
            print("==========================================STARTING Search=====================================================")


            # global workbook_Title
            workbook_Title = globalVars.page3File
            print(workbook_Title)
            # global workbook
            workbook = load_workbook(filename=workbook_Title)
            # global sheet
            sheet = workbook.active

            rowIndex=2
            columnIndex=1
            lastNameColumnIndex = columnIndex+1
            ssnColumnIndex = columnIndex+2
            dodColumnIndex = columnIndex+3
            flightColumnIndex = columnIndex+4
            signInColumnIndex = columnIndex+5
            signOutColumnIndex = columnIndex+6
            found=False
            firstNamecellref = sheet.cell(row=rowIndex,column=columnIndex)
            lastNamecellref = sheet.cell(row=rowIndex,column=lastNameColumnIndex)
            ssncellref = sheet.cell(row=rowIndex,column=ssnColumnIndex)
            dodCellref = sheet.cell(row=rowIndex,column=dodColumnIndex)
            flightCellRef = sheet.cell(row=rowIndex,column=flightColumnIndex)
            signInCellRef = sheet.cell(row=rowIndex,column=signInColumnIndex)
            signOutCellRef = sheet.cell(row=rowIndex,column=signOutColumnIndex)
            signOutMonth = str(signOutCellRef.value)[0:2]
            signOutMonth = int(signOutMonth)
            signOutDay = str(signOutCellRef.value)[3:5]
            signOutDay = int(signOutDay)
            signOutYear = str(signOutCellRef.value)[6:10]
            signOutYear = int(signOutYear)
            print("SIGN OUT YEAR: "+str(signOutYear))
            currentMonth = str(globalVars.signOutDate)[0:2]
            currentMonth = int(currentMonth)
            currentDay = str(globalVars.signOutDate)[3:5]
            currentDay = int(currentDay)
            currentYear = str(globalVars.signOutDate)[6:10]
            print("CURRENT YEAR: "+currentYear)
            currentYear = int(currentYear)
            print("CURRENT YEAR: "+str(currentYear))


            while firstNamecellref.value:
                signOutMonth = str(signOutCellRef.value)[0:2]
                signOutMonth = int(signOutMonth)
                signOutDay = str(signOutCellRef.value)[3:5]
                signOutDay = int(signOutDay)
                signOutYear = str(signOutCellRef.value)[6:10]
                signOutYear = int(signOutYear)
                print("SIGN OUT: "+str(signOutCellRef.value))
                print("CURRENT MONTH: "+str(currentMonth))
                print("CURRENT DAY: "+str(currentDay))
                print("SIGNOUT MONTH: "+str(signOutMonth))
                print("SIGNOUT Day: "+str(signOutDay))
                if signOutMonth == currentMonth and signOutYear == currentYear and signOutDay <= currentDay+2:
                    # firstNamecellref = sheet.cell(rowIndex,column=columnIndex+1)
                    globalVars.returnedFirstName = str(firstNamecellref.value)
                    globalVars.returnedLastName = str(lastNamecellref.value)
                    globalVars.returnedSsn = str(ssncellref.value)
                    # print("SSN: "+str(globalVars.returnedSsn))
                    globalVars.returnedDod = str(dodCellref.value)
                    globalVars.returnedFlight = str(flightCellRef.value)
                    globalVars.returnedSignIn = str(signInCellRef.value)
                    globalVars.returnedSignOut = str(signOutCellRef.value)
                    self.found.emit()
                    found =True

                    # time.sleep(1)
                    rowIndex+=1
                else:
                    rowIndex+=1
                firstNamecellref = sheet.cell(row=rowIndex,column=columnIndex)
                lastNamecellref = sheet.cell(row=rowIndex,column=lastNameColumnIndex)
                ssncellref = sheet.cell(row=rowIndex,column=ssnColumnIndex)
                dodCellref = sheet.cell(row=rowIndex,column=dodColumnIndex)
                flightCellRef = sheet.cell(row=rowIndex,column=flightColumnIndex)
                signInCellRef = sheet.cell(row=rowIndex,column=signInColumnIndex)
                signOutCellRef = sheet.cell(row=rowIndex,column=signOutColumnIndex)
                print("SIGN OUT: "+str(signInCellRef.value))
                print(firstNamecellref.coordinate)
            self.finished.emit()

  
        except():
            print(traceback.format_exc())
            # messagebox.showwarning(title="Error Occured", message="something went wrong in EXISTING BOOK. Check your entries and try again")
            if path.exists("errors.txt") == True:
                ct = datetime.datetime.now() 
                with open("errors.txt", "a") as file:
                    file.write("\n"+str(ct)+"\n"+str(traceback.format_exc()))
            else:
                ct = datetime.datetime.now() 
                with open("errors.txt", "x") as file:
                    file.write("\n"+str(ct)+"\n"+str(traceback.format_exc()))
  

    