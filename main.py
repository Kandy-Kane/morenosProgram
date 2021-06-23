# from tab6 import Worker3
# from newBook import Worker
from singleEntry import Worker2
from search import Worker3
from checkSignOuts import Worker4
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.cell import cell
from openpyxl.cell.cell import ERROR_CODES
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, alignment
from openpyxl.styles.fills import fills
from openpyxl.worksheet.dimensions import SheetDimension
from openpyxl.utils import get_column_letter
import sys
import os
from PyQt5 import QtCore, QtGui,QtWidgets,uic
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from PyQt5.QtCore import QObject, QThread, pyqtSignal
from PyQt5.QtCore import QPropertyAnimation, QPoint
from os import path
from PyQt5.QtGui import QMovie
import logging
import threading
import time
import globalVars
import pyautogui
import subprocess
from random import randint
import datetime
import traceback





class AppDemo(QMainWindow):
    def __init__(self):
        super().__init__()
        uic.loadUi('ver1M.ui',self)
        # global finishedLabel
        # global failedLabel
        global statusWidget
        statusWidget = self.stackedWidget_2
        global movie
        self.movie = QMovie("./images/spinning.gif")
        self.label_10.setMovie(self.movie)
        self.movie.start()
        global movie2
        self.movie2 = QMovie("./images/loading4.gif")
        self.label_6.setMovie(self.movie2)
        self.movie2.start()
        self.stackedWidget_2.setCurrentWidget(self.page_8)
        self.singleEntryButton.clicked.connect(self.showPage1)
        self.newWorkbookButton.clicked.connect(self.showPage2)
        # self.dateTimeEdit.setDateTime(QtCore.QDateTime.currentDateTime())
        self.readFromFileButton.clicked.connect(self.showPage3)
        self.addFullQualButton.clicked.connect(self.showPage4)
        self.addMirsButton.clicked.connect(self.daKinter)
        
        # realmScroll.setWidget(layout.widget())


        
        # self.scrollArea.setWidgetResizable(False)
        self.scrollArea_4.widgetResizable()
        self.scrollArea_4.setWidgetResizable(True)
        # layout = QVBoxLayout(self.scrollArea_3)
        
        # datetime = QDateTime.currentDateTime()
        # # datetime.setDisplayFormat("dd/MM/yy hh:mm")
        # text = datetime.toString()
        # self.timeLabel.setText(text)
        self.toolButton.clicked.connect(self.launchDialog)
        self.toolButton_5.clicked.connect(self.launchDialog2)
        # statusWidget.setHidden(True)
        # self.page2Submit.clicked.connect(self)
        self.page1Submit.clicked.connect(self.singleEntry)
        self.stackedWidget_2.setCurrentWidget(self.page_9)
        
        self.thumbListWidget.setSortingEnabled(True)
        # self.populateButton.clicked.connect(self.daKinter)
        # self.moveButton.clicked.connect(self.select)
        self.toggleFrame.setHidden(True)
        self.hideToggle.clicked.connect(self.showToggle)
        self.toggleButton.clicked.connect(self.showToggle)
        # self.progressBar.setValue(globalVars.progressbarValue)
        self.progressBar.setHidden(True)
        self.filterWidget.setHidden(True)
        self.tabWidget.setHidden(True)
        self.settingsButton.clicked.connect(self.showSettings)
        # self.testButton.clicked.connect(self.newReportProgress)

        self.opacity_effect = QGraphicsOpacityEffect()
        self.opacity_effect.setOpacity(0.7)
        self.filterWidget.setGraphicsEffect(self.opacity_effect)

        self.opacity_effect2 = QGraphicsOpacityEffect()
        self.opacity_effect2.setOpacity(1)
        self.tabWidget.setGraphicsEffect(self.opacity_effect2)
        self.toolButton_3.clicked.connect(self.launchDialog3)
        self.page2Submit_1.clicked.connect(self.newNameSearch)
        self.page2Submit_2.clicked.connect(self.newSsnSearch)
        self.page2Submit_3.clicked.connect(self.newDodSearch)
        self.page2Submit_4.clicked.connect(self.newFlightSearch)
        self.page3Submit_1.clicked.connect(self.checkSignOuts)
        # self.clearButton.clicked.connect(self.clearProgressFrame)
        self.radioButton.toggled.connect(lambda:self.btnstate(self.radioButton))
        self.radioButton_2.toggled.connect(lambda:self.btnstate(self.radioButton_2))
        self.radioButton_3.toggled.connect(lambda:self.btnstate(self.radioButton_3))
        self.timer = QTimer()
        # self.formLayout.setVerticalSpacing(30)

        


        # global mainemployeeslist
        # with open('./employees.txt') as f:
        #     mainemployeeslist = [line.rstrip() for line in f]
        # for item in mainemployeeslist:
        #     self.thumbListWidget.insertItem(0,item)

        #DATE AND TIME STUFF
        timer = QTimer(self)
        timer.timeout.connect(self.showtime)
        timer.start()




    def clearFrame(self):
        self.progressFrame = self.layout
        while self.layout.count():
            child = self.layout.takeAt(0)
            if child.widget() is not None:
                child.widget().deleteLater()
            elif child.layout() is not None:
                self.clearLayout(child.layout())

    def btnstate(self,b):
        if self.radioButton.isChecked() == True:
            self.sideBar.setStyleSheet('background-color: rgb(0, 150, 225);')
            self.topbar.setStyleSheet('background-color: rgb(117, 117, 117);')
            self.bottombar.setStyleSheet('background-color: rgb(117, 117, 117);')
            self.page.setStyleSheet('background-color: rgb(232, 232, 232);')
            self.singleEntryButton.setStyleSheet('background-color: rgb(0, 150, 225);')
            self.singleEntryButton.setStyleSheet("QPushButton::hover""{""background-color :rgb(255, 62, 65);""}")
            self.newWorkbookButton.setStyleSheet('background-color: rgb(0, 150, 225);')
            self.newWorkbookButton.setStyleSheet("QPushButton::hover""{""background-color :rgb(255, 62, 65);""}")
            self.readFromFileButton.setStyleSheet('background-color: rgb(0, 150, 225);')
            self.readFromFileButton.setStyleSheet("QPushButton::hover""{""background-color :rgb(255, 62, 65);""}")
            self.addFullQualButton.setStyleSheet('background-color: rgb(0, 150, 225);')
            self.addFullQualButton.setStyleSheet("QPushButton::hover""{""background-color :rgb(255, 62, 65);""}")
            self.addMirsButton.setStyleSheet('background-color: rgb(0, 150, 225);')
            self.addMirsButton.setStyleSheet("QPushButton::hover""{""background-color :rgb(255, 62, 65);""}")
            self.page_2.setStyleSheet('background-color: rgb(232, 232, 232);')
            self.page_3.setStyleSheet('background-color: rgb(232, 232, 232);')
            self.page_4.setStyleSheet('background-color: rgb(232, 232, 232);')
        elif self.radioButton_2.isChecked() == True:
            self.sideBar.setStyleSheet('background-color: rgb(83, 0, 89);')
            self.topbar.setStyleSheet('background-color: rgb(34, 34, 34);')
            self.bottombar.setStyleSheet('background-color: rgb(34, 34, 34);')
            self.page.setStyleSheet('background-color: black;')
            self.singleEntryButton.setStyleSheet('background-color: rgb(83, 0, 89);')
            self.singleEntryButton.setStyleSheet("QPushButton::hover""{""background-color: rgb(184, 81, 174);""}")
            self.newWorkbookButton.setStyleSheet('background-color: rgb(83, 0, 89);')
            self.newWorkbookButton.setStyleSheet("QPushButton::hover""{""background-color: rgb(184, 81, 174);""}")
            self.readFromFileButton.setStyleSheet('background-color: rgb(83, 0, 89);')
            self.readFromFileButton.setStyleSheet("QPushButton::hover""{""background-color: rgb(184, 81, 174);""}")
            self.addFullQualButton.setStyleSheet('background-color: rgb(83, 0, 89);')
            self.addFullQualButton.setStyleSheet("QPushButton::hover""{""background-color: rgb(184, 81, 174);""}")
            self.addMirsButton.setStyleSheet('background-color: rgb(83, 0, 89);')
            self.addMirsButton.setStyleSheet("QPushButton::hover""{""background-color: rgb(184, 81, 174);""}")
            self.page_2.setStyleSheet('background-color: black;')
            self.page_3.setStyleSheet('background-color: black;')
            self.page_4.setStyleSheet('background-color: black;')
        elif self.radioButton_3.isChecked() == True:
            self.sideBar.setStyleSheet('background-color: rgb(111, 0, 0);')
            self.topbar.setStyleSheet('background-color: rgb(34, 34, 34);')
            self.bottombar.setStyleSheet('background-color: rgb(34, 34, 34);')
            self.page.setStyleSheet('background-color: rgb(0, 79, 91);')
            self.singleEntryButton.setStyleSheet('background-color: rgb(111, 0, 0);')
            self.singleEntryButton.setStyleSheet("QPushButton::hover""{""background-color: rgb(0, 255, 247);""}")
            self.newWorkbookButton.setStyleSheet('background-color: rgb(111, 0, 0);')
            self.newWorkbookButton.setStyleSheet("QPushButton::hover""{""background-color: rgb(0, 255, 247);""}")
            self.readFromFileButton.setStyleSheet('background-color: rgb(111, 0, 0);')
            self.readFromFileButton.setStyleSheet("QPushButton::hover""{""background-color: rgb(0, 255, 247);""}")
            self.addFullQualButton.setStyleSheet('background-color: rgb(111, 0, 0);')
            self.addFullQualButton.setStyleSheet("QPushButton::hover""{""background-color: rgb(0, 255, 247);""}")
            self.addMirsButton.setStyleSheet('background-color: rgb(111, 0, 0);')
            self.addMirsButton.setStyleSheet("QPushButton::hover""{""background-color: rgb(0, 255, 247);""}")
            self.page_2.setStyleSheet('background-color: rgb(0, 79, 91);')
            self.page_3.setStyleSheet('background-color: rgb(0, 79, 91);')
            self.page_4.setStyleSheet('background-color: rgb(0, 79, 91);')
            
        
        # self.sideBar.setStyleSheet('background-color: black;')
        




    def showtime(self):
        datetime = QDate.currentDate()
        text = datetime.toString('MM/dd/yyyy')
        
        self.timeLabel.setText(text)

        
    def showToggle(self, checked):
        if self.toggleFrame.isHidden():
            self.toggleFrame.setHidden(False)
            self.filterWidget.setHidden(False)
            
        else:
            self.toggleFrame.setHidden(True)
            self.filterWidget.setHidden(True)
            self.tabWidget.setHidden(True)
            

    def showSettings(self):
        #skittle
        if self.tabWidget.isHidden():
            self.tabWidget.setHidden(False)
            self.anim = QPropertyAnimation(self.tabWidget, b"pos")
            effect = QGraphicsOpacityEffect(self.tabWidget)
            self.tabWidget.setGraphicsEffect(effect)
            self.anim.setStartValue(QPoint(240, 35))
            self.anim.setEndValue(QPoint(240,45 ))
            self.anim.setDuration(250)
            self.anim_2 = QPropertyAnimation(effect, b"opacity")
            self.anim_2.setStartValue(0)
            self.anim_2.setEndValue(1)
            self.anim_2.setDuration(500)
            self.anim_group = QParallelAnimationGroup()
            self.anim_group.addAnimation(self.anim)
            self.anim_group.addAnimation(self.anim_2)
            self.anim_group.start()
            
        else:
            self.tabWidget.setHidden(True)
            # self.filterWidget.setHidden(True)

      
        

#PAGE1==========================================SINGLE ENTRY=============================================
    # global initY
    # initY = 8
    # global namei
    # namei = 0
    def newFound(self):
        # self.spacer = QWidget()
        # self.spacer.setBaseSize(300,150)
        # self.spacer.setAutoFillBackground(True)
        # self.spacer.setStyleSheet('background-color: red;')
        global namei
        global initY
        i = 35
        firstNameLabel = QLabel('First Name: '+str(globalVars.returnedFirstName))
        # firstNameLabel.setMinimumHeight(50)
        


        lastNameLabel = QLabel('Last Name: '+str(globalVars.returnedLastName))
        # lastNameLabel.setMinimumHeight(50)


        ssnLabel = QLabel('SSN: '+str(globalVars.returnedSsn))
        # ssnLabel.setMinimumHeight(50)
        self.formLayout.addWidget(firstNameLabel)
        self.formLayout.addWidget(lastNameLabel)
        self.formLayout.addWidget(ssnLabel)
        # self.formLayout.addWidget(QLabel('First Name: '+str(globalVars.returnedFirstName)))
        # self.formLayout.addWidget(QLabel('Last Name: '+str(globalVars.returnedLastName)))
        # self.formLayout.addWidget(QLabel('SSN: '+str(globalVars.returnedSsn)))
        self.formLayout.addWidget(QLabel('DOD: '+str(globalVars.returnedDod)))
        self.formLayout.addWidget(QLabel('Flight: '+str(globalVars.returnedFlight)))
        self.formLayout.addWidget(QLabel('Sign In Date: '+str(globalVars.returnedSignIn)))
        self.formLayout.addWidget(QLabel('Sign Out Date: '+str(globalVars.returnedSignOut)))
        self.formLayout.addWidget(QLabel('================================================'))
       
        
        

        # self.dod = QLabel(self.scrollAreaWidgetContents_3)
        # self.dod.setText('DOD: '+str(globalVars.returnedDod))
        # self.dod.setGeometry(10,initY+i*3,250,30)
        # self.dod.show()

        # self.flight = QLabel(self.scrollAreaWidgetContents_3)
        # self.flight.setText('Flight: '+str(globalVars.returnedFlight))
        # self.flight.setGeometry(10,initY+i*4,250,30)
        # self.flight.show()

        # self.signIn = QLabel(self.scrollAreaWidgetContents_3)
        # self.signIn.setText('Sign In Date: '+str(globalVars.returnedSignIn))
        # self.signIn.setGeometry(10,initY+i*5,350,30)
        # self.signIn.show()

        # self.signOut = QLabel(self.scrollAreaWidgetContents_3)
        # self.signOut.setText('Sign Out Date: '+str(globalVars.returnedSignOut))
        # self.signOut.setGeometry(10,initY+i*6,350,30)
        # self.signOut.show()
        
        # self.progress.setValue(globalVars.progressbarValue)
        
        # initY+=300

    
    # def clearProgressFrame(self):
    #         self.progress.deleteLater()
    #         self.progress.deleteLater()

    # def page4ReportProgress(self):
    #     # self.progress.setHidden(False)
    #     self.progress.setValue(globalVars.progressbarValue)
    #     # animation = QPropertyAnimation(self.progressBar, "value")
    #     # animation.setDuration(2000)
    #     # animation.setStartValue(0)
    #     # animation.setEndValue(100)
    #     # animation.start()


    # def reportProgress(self):
    #     self.progressBar.setHidden(False)
    #     self.progressBar.setValue(globalVars.progressbarValue)
    #     # animation = QPropertyAnimation(self.progressBar, "value")
    #     # animation.setDuration(2000)
    #     # animation.setStartValue(0)
    #     # animation.setEndValue(100)
    #     # animation.start()





    def newFound2(self):
        global namei
        global initY
        i = 35
        firstNameLabel = QLabel('First Name: '+str(globalVars.returnedFirstName))
        # firstNameLabel.setMinimumHeight(50)
        


        lastNameLabel = QLabel('Last Name: '+str(globalVars.returnedLastName))
        # lastNameLabel.setMinimumHeight(50)


        ssnLabel = QLabel('SSN: '+str(globalVars.returnedSsn))
        # ssnLabel.setMinimumHeight(50)
        self.formLayout_2.addWidget(firstNameLabel)
        self.formLayout_2.addWidget(lastNameLabel)
        self.formLayout_2.addWidget(ssnLabel)
        # self.formLayout.addWidget(QLabel('First Name: '+str(globalVars.returnedFirstName)))
        # self.formLayout.addWidget(QLabel('Last Name: '+str(globalVars.returnedLastName)))
        # self.formLayout.addWidget(QLabel('SSN: '+str(globalVars.returnedSsn)))
        self.formLayout_2.addWidget(QLabel('DOD: '+str(globalVars.returnedDod)))
        self.formLayout_2.addWidget(QLabel('Flight: '+str(globalVars.returnedFlight)))
        self.formLayout_2.addWidget(QLabel('Sign In Date: '+str(globalVars.returnedSignIn)))
        self.formLayout_2.addWidget(QLabel('Sign Out Date: '+str(globalVars.returnedSignOut)))
        self.formLayout_2.addWidget(QLabel('================================================'))
       

    
    # def clearProgressFrame(self):
    #         self.progress.deleteLater()
    #         self.progress.deleteLater()

    # def page4ReportProgress(self):
    #     # self.progress.setHidden(False)
    #     self.progress.setValue(globalVars.progressbarValue)
    #     # animation = QPropertyAnimation(self.progressBar, "value")
    #     # animation.setDuration(2000)
    #     # animation.setStartValue(0)
    #     # animation.setEndValue(100)
    #     # animation.start()


    # def reportProgress(self):
    #     self.progressBar.setHidden(False)
    #     self.progressBar.setValue(globalVars.progressbarValue)
    #     # animation = QPropertyAnimation(self.progressBar, "value")
    #     # animation.setDuration(2000)
    #     # animation.setStartValue(0)
    #     # animation.setEndValue(100)
    #     # animation.start()




#==============================================SINGLE ENTRY=====================================================#
    def singleEntry(self):
        # statusWidget.setVisible(True)
        # self.label_6.setVisible(True)

        #E1
        # print(globalVars.page1e1)
        globalVars.page1File = self.fileNameLine.text()
        globalVars.page1File = str(globalVars.page1File)
        print(globalVars.page1File)

        #e2
        # print(globalVars.page1e2)
        globalVars.firstName = self.fNameEntry.text()
        globalVars.firstName = str(globalVars.firstName).upper()
        print(globalVars.firstName)

        #e3
        # print(globalVars.page1e3)
        globalVars.lastName = self.lNameEntry.text()
        globalVars.lastName = str(globalVars.lastName).upper()
        print(globalVars.lastName)

         # print(globalVars.page1e3)
        globalVars.ssn = self.ssnEntry.text()
        globalVars.ssn = str(globalVars.ssn)
        print(globalVars.ssn)


         # print(globalVars.page1e3)
        globalVars.dod = self.dodEntry.text()
        globalVars.dod = str(globalVars.dod)
        print(globalVars.dod)

        globalVars.flight = self.flightEntry.text()
        globalVars.flight = str(globalVars.flight).upper()
        print(globalVars.flight)

        globalVars.signInDate = self.page1StartDate.text()
        globalVars.signInDate = str(globalVars.signInDate)
        print(globalVars.signInDate)

        globalVars.signOutDate = self.page1EndDate.text()
        globalVars.signOutDate = str(globalVars.signOutDate)
        print(globalVars.signOutDate)

        if globalVars.page1File == '':
            self.stackedWidget_2.setCurrentWidget(self.page_12)
            return

        if globalVars.firstName == '':
            self.stackedWidget_2.setCurrentWidget(self.page_12)
            return
        if globalVars.lastName == '':
            self.stackedWidget_2.setCurrentWidget(self.page_12)
            return
        if globalVars.ssn == '':
            self.stackedWidget_2.setCurrentWidget(self.page_12)
            return
        if globalVars.dod == '':
            self.stackedWidget_2.setCurrentWidget(self.page_12)
            return
        if globalVars.flight == '':
            self.stackedWidget_2.setCurrentWidget(self.page_12)
            return

        if path.exists(str(globalVars.page1File)) != True:
            self.stackedWidget_2.setCurrentWidget(self.page_7)
            # statusWidget.setHidden(False)
            return
        else:
            # self.isHoliday.setChecked(False)
            # self.label_6.setHidden(False)
            print("FIRST CHECK: ")
            print(statusWidget.isVisible())
            print("FIRST CHECK: ")
            print(self.label_6.isVisible())
            self.stackedWidget_2.setCurrentWidget(self.page_8)
            statusWidget.setHidden(False)
            # self.label_6.setVisible(True)
            global movie2
            self.movie2.start()
            # Worker2.addEntry(self)
            self.thread = QThread()
            # Step 3: Create a worker object
            self.worker = Worker2()
            # Step 4: Move worker to the thread
            self.worker.moveToThread(self.thread)
            # Step 5: Connect signals and slots
            self.thread.started.connect(self.worker.addEntry)
            self.worker.finished.connect(self.thread.quit)
            self.worker.finished.connect(self.worker.deleteLater)
            self.thread.finished.connect(self.thread.deleteLater)
            # self.worker.progress.connect(self.reportProgress)
            # Step 6: Start the thread
            self.thread.start()
            self.thread.finished.connect(self.runFinished)

#PAGE4=================================================================================================


    def page4Error(self,error):
        print("error is running")
        self.msg = QMessageBox()
        self.msg.setIcon(QMessageBox.Critical)
        self.msg.setText("Error When Adding Qual")
        if error == 'special':
            self.msg.setInformativeText("It seems one of your entries has special characters in it. Please check your entries.")
        elif error == 'nostartdate':
            self.msg.setInformativeText("It seems one of the Start Date cells in your read From file is empty. Please check your entries.")
        elif error == 'general':
            self.msg.setInformativeText("An unknown general error occured when adding full Qual. Please check your entries.")
        elif error == 'readFrom':
            self.msg.setInformativeText("An unknown general error occured when reading your read From file. Please check your entries.")
        elif error == 'instructorOnLeave':
            self.msg.setIcon(QMessageBox.Information)
            self.msg.setInformativeText("And instructor is on leave for one of the dates you scheduled them for.")
            self.msg.setWindowTitle("INSTRUCTOR ON LEAVE")
        self.msg.setWindowTitle("FULL ADD QUAL ERROR")
        # self.msg.setDetailedText("The details are as follows:")
        self.msg.show()
        

        
        




    def newNameSearch(self):
        for i in reversed(range(self.formLayout.count())): 
            self.formLayout.itemAt(i).widget().deleteLater()
        print("=============================================START=======================================================")
        # checkBox = self.checkBox
        globalVars.firstName = self.fNameEntry_2.text()
        globalVars.lastName = self.lNameEntry_2.text()
        globalVars.page2File = self.fileNameLine_3.text()
        # statusWidget.setVisible(True)
        # self.label_6.setVisible(True)
        globalVars.page2File = str(globalVars.page2File)
        print(globalVars.page2File)

        if globalVars.firstName == '':
            self.stackedWidget_2.setCurrentWidget(self.page_12)
            return
        if globalVars.lastName == '':
            self.stackedWidget_2.setCurrentWidget(self.page_12)
            return
        


        
        if path.exists(str(globalVars.page2File)) != True:
            self.stackedWidget_2.setCurrentWidget(self.page_10)
            # statusWidget.setHidden(False)
            return
        else:
            self.stackedWidget_2.setCurrentWidget(self.page_8)
            # Worker3.addFullQual(self)
            statusWidget.setHidden(False)
            # self.label_6.setVisible(True)
            global movie2
            self.movie2.start()
            # Worker3.search()
            self.thread = QThread()
            # Step 3: Create a worker object
            self.worker = Worker3()
            # Step 4: Move worker to the thread
            self.worker.moveToThread(self.thread)
            self.thread.started.connect(self.worker.search)
            self.worker.finished.connect(self.thread.quit)
            self.worker.finished.connect(self.worker.deleteLater)
            self.thread.finished.connect(self.thread.deleteLater)
            self.worker.found.connect(self.newFound)
            # self.worker.progress.connect(self.page4ReportProgress)
            # self.worker.error.connect(self.page4Error)
            # Step 6: Start the thread
            self.thread.start()
            self.thread.finished.connect(self.runFinished)



    def newSsnSearch(self):
        for i in reversed(range(self.formLayout.count())): 
            self.formLayout.itemAt(i).widget().deleteLater()
        print("=============================================START=======================================================")
        # checkBox = self.checkBox
        globalVars.ssn = self.ssnEntry_2.text()
        globalVars.ssn=str(globalVars.ssn)
        globalVars.page2File = self.fileNameLine_3.text()
        # statusWidget.setVisible(True)
        # self.label_6.setVisible(True)
        globalVars.page2File = str(globalVars.page2File)
        print(globalVars.page2File)

        if globalVars.ssn == '':
            self.stackedWidget_2.setCurrentWidget(self.page_12)
            return

        
        if path.exists(str(globalVars.page2File)) != True:
            self.stackedWidget_2.setCurrentWidget(self.page_10)
            # statusWidget.setHidden(False)
            return
        else:
            self.stackedWidget_2.setCurrentWidget(self.page_8)
            # Worker3.addFullQual(self)
            statusWidget.setHidden(False)
            # self.label_6.setVisible(True)
            global movie2
            self.movie2.start()
            # Worker3.search()
            self.thread = QThread()
            # Step 3: Create a worker object
            self.worker = Worker3()
            # Step 4: Move worker to the thread
            self.worker.moveToThread(self.thread)
            self.thread.started.connect(self.worker.ssnSearch)
            self.worker.finished.connect(self.thread.quit)
            self.worker.finished.connect(self.worker.deleteLater)
            self.thread.finished.connect(self.thread.deleteLater)
            self.worker.found.connect(self.newFound)
            # self.worker.progress.connect(self.page4ReportProgress)
            # self.worker.error.connect(self.page4Error)
            # Step 6: Start the thread
            self.thread.start()
            self.thread.finished.connect(self.runFinished)


    def newDodSearch(self):
        for i in reversed(range(self.formLayout.count())): 
            self.formLayout.itemAt(i).widget().deleteLater()
        print("=============================================START=======================================================")
        # checkBox = self.checkBox
        globalVars.dod = self.dodEntry_2.text()
        globalVars.dod=str(globalVars.dod)
        globalVars.page2File = self.fileNameLine_3.text()
        # statusWidget.setVisible(True)
        # self.label_6.setVisible(True)
        globalVars.page2File = str(globalVars.page2File)
        print(globalVars.page2File)

        if globalVars.dod == '':
            self.stackedWidget_2.setCurrentWidget(self.page_12)
            return

        
        if path.exists(str(globalVars.page2File)) != True:
            self.stackedWidget_2.setCurrentWidget(self.page_10)
            # statusWidget.setHidden(False)
            return
        else:
            self.stackedWidget_2.setCurrentWidget(self.page_8)
            # Worker3.addFullQual(self)
            statusWidget.setHidden(False)
            # self.label_6.setVisible(True)
            global movie2
            self.movie2.start()
            # Worker3.search()
            self.thread = QThread()
            # Step 3: Create a worker object
            self.worker = Worker3()
            # Step 4: Move worker to the thread
            self.worker.moveToThread(self.thread)
            self.thread.started.connect(self.worker.dodSearch)
            self.worker.finished.connect(self.thread.quit)
            self.worker.finished.connect(self.worker.deleteLater)
            self.thread.finished.connect(self.thread.deleteLater)
            self.worker.found.connect(self.newFound)
            # self.worker.progress.connect(self.page4ReportProgress)
            # self.worker.error.connect(self.page4Error)
            # Step 6: Start the thread
            self.thread.start()
            self.thread.finished.connect(self.runFinished)



    def newFlightSearch(self):
        for i in reversed(range(self.formLayout.count())): 
            self.formLayout.itemAt(i).widget().deleteLater()
        print("=============================================START=======================================================")
        # checkBox = self.checkBox
        globalVars.flight = self.flightEntry_2.text()
        globalVars.flight=str(globalVars.flight)
        globalVars.page2File = self.fileNameLine_3.text()
        # statusWidget.setVisible(True)
        # self.label_6.setVisible(True)
        globalVars.page2File = str(globalVars.page2File)
        print(globalVars.page2File)

        if globalVars.flight == '':
            self.stackedWidget_2.setCurrentWidget(self.page_12)
            return

        
        if path.exists(str(globalVars.page2File)) != True:
            self.stackedWidget_2.setCurrentWidget(self.page_10)
            # statusWidget.setHidden(False)
            return
        else:
            self.stackedWidget_2.setCurrentWidget(self.page_8)
            # Worker3.addFullQual(self)
            statusWidget.setHidden(False)
            # self.label_6.setVisible(True)
            global movie2
            self.movie2.start()
            # Worker3.search()
            self.thread = QThread()
            # Step 3: Create a worker object
            self.worker = Worker3()
            # Step 4: Move worker to the thread
            self.worker.moveToThread(self.thread)
            self.thread.started.connect(self.worker.flightSearch)
            self.worker.finished.connect(self.thread.quit)
            self.worker.finished.connect(self.worker.deleteLater)
            self.thread.finished.connect(self.thread.deleteLater)
            self.worker.found.connect(self.newFound)
            # self.worker.progress.connect(self.page4ReportProgress)
            # self.worker.error.connect(self.page4Error)
            # Step 6: Start the thread
            self.thread.start()
            self.thread.finished.connect(self.runFinished)



    def checkSignOuts(self):
        for i in reversed(range(self.formLayout_2.count())): 
            self.formLayout_2.itemAt(i).widget().deleteLater()
        print("=============================================START=======================================================")
        # checkBox = self.checkBox
        globalVars.signOutDate = str(self.timeLabel.text())
        print("GLOBAL SIGN OUT: "+str(globalVars.signOutDate))
        globalVars.signOutDate=str(globalVars.signOutDate)
        globalVars.page3File = self.fileNameLine_4.text()
        # statusWidget.setVisible(True)
        # self.label_6.setVisible(True)
        globalVars.page3File = str(globalVars.page3File)
        print(globalVars.page3File)

        
        if path.exists(str(globalVars.page3File)) != True:
            self.stackedWidget_2.setCurrentWidget(self.page_10)
            # statusWidget.setHidden(False)
            return
        else:
            self.stackedWidget_2.setCurrentWidget(self.page_8)
            # Worker3.addFullQual(self)
            statusWidget.setHidden(False)
            # self.label_6.setVisible(True)
            global movie2
            self.movie2.start()
            # Worker3.search()
            self.thread = QThread()
            # Step 3: Create a worker object
            self.worker = Worker4()
            # Step 4: Move worker to the thread
            self.worker.moveToThread(self.thread)
            self.thread.started.connect(self.worker.signOutSearch)
            self.worker.finished.connect(self.thread.quit)
            self.worker.finished.connect(self.worker.deleteLater)
            self.thread.finished.connect(self.thread.deleteLater)
            self.worker.found.connect(self.newFound2)
            # self.worker.progress.connect(self.page4ReportProgress)
            # self.worker.error.connect(self.page4Error)
            # Step 6: Start the thread
            self.thread.start()
            self.thread.finished.connect(self.runFinished)




    




#PAGE5======================================ADD MIRS======================================================
    def daKinter(self):
        subprocess.Popen(['python' , 'ver8.py'])

#EXTRAS===================================================================================================


            
    def runFinished(self):
        globalVars.progressbarValue=0
        self.progressBar.setHidden(True)
        print("Second CHECK: ")
        print(statusWidget.isVisible())
        print("Second CHECK: ")
        print(self.label_6.isVisible())
        self.stackedWidget_2.setCurrentWidget(self.page_6)
        # statusWidget.setHidden(False)
        # effect = QGraphicsOpacityEffect(self.label_19)
        # self.label_19.setGraphicsEffect(effect)
        # self.anim_2 = QPropertyAnimation(effect, b"opacity")
        # self.anim_2.setStartValue(1)
        # self.anim_2.setEndValue(0)
        # self.anim_2.setDuration(5000)
        # self.anim_2.start()
        # self.fileNamePage2.setText("")
        self.movie2.stop()
        
        

    def launchDialog(self):
        file_Filter = 'Data File (*.xlsx *.csv *.dat);; Excel File (*.xlsx *.xls)'
        filePath,_ = QFileDialog.getOpenFileName(
            parent=self,
            caption="Select File",
            directory=os.getcwd(),
            filter=file_Filter,
            initialFilter='Excel File (*.xlsx *.xls)'
        )
        # fileName = QFileInfo(filePath).fileName()
        url = QUrl.fromLocalFile(filePath)
        global fileName
        fileName = str(url.fileName())
        print(url.fileName())
        self.fileNameLine.setText(fileName)

    def launchDialog2(self):
        file_Filter = 'Data File (*.xlsx *.csv *.dat);; Excel File (*.xlsx *.xls)'
        filePath,_ = QFileDialog.getOpenFileName(
            parent=self,
            caption="Select File",
            directory=os.getcwd(),
            filter=file_Filter,
            initialFilter='Excel File (*.xlsx *.xls)'
        )
        # fileName = QFileInfo(filePath).fileName()
        url = QUrl.fromLocalFile(filePath)
        global fileName
        fileName = str(url.fileName())
        print(url.fileName())
        self.fileNameLine_3.setText(fileName)


    def launchDialog3(self):
        file_Filter = 'Data File (*.xlsx *.csv *.dat);; Excel File (*.xlsx *.xls)'
        filePath,_ = QFileDialog.getOpenFileName(
            parent=self,
            caption="Select File",
            directory=os.getcwd(),
            filter=file_Filter,
            initialFilter='Excel File (*.xlsx *.xls)'
        )
        # fileName = QFileInfo(filePath).fileName()
        url = QUrl.fromLocalFile(filePath)
        global fileName
        fileName = str(url.fileName())
        print(url.fileName())
        self.fileNameLine_4.setText(fileName)
        

    

        
       

    def showPage1(self):
        print("Page1")
        self.anim = QPropertyAnimation(self.stackedWidget, b"pos")
        effect = QGraphicsOpacityEffect(self.stackedWidget)
        self.stackedWidget.setGraphicsEffect(effect)
        self.anim.setStartValue(QPoint(200, 180))
        self.anim.setEndValue(QPoint(200, 150))
        self.anim.setDuration(500)
        self.anim_2 = QPropertyAnimation(effect, b"opacity")
        self.anim_2.setStartValue(0)
        self.anim_2.setEndValue(1)
        self.anim_2.setDuration(500)
        self.anim_group = QParallelAnimationGroup()
        self.anim_group.addAnimation(self.anim)
        self.anim_group.addAnimation(self.anim_2)
        self.anim_group.start()
        self.stackedWidget.setCurrentWidget(self.page)


    def showPage2(self):
        print("Page2")
        self.anim = QPropertyAnimation(self.stackedWidget, b"pos")
        effect = QGraphicsOpacityEffect(self.stackedWidget)
        self.stackedWidget.setGraphicsEffect(effect)
        self.anim.setStartValue(QPoint(200, 100))
        self.anim.setEndValue(QPoint(200, 150))
        self.anim.setEasingCurve(QEasingCurve.OutBounce)
        self.anim.setDuration(500)
        self.anim_2 = QPropertyAnimation(effect, b"opacity")
        self.anim_2.setStartValue(0)
        self.anim_2.setEndValue(1)
        self.anim_2.setDuration(500)
        self.anim_group = QParallelAnimationGroup()
        self.anim_group.addAnimation(self.anim)
        self.anim_group.addAnimation(self.anim_2)
        self.anim_group.start()
        self.stackedWidget.setCurrentWidget(self.page_2)

    def showPage3(self):
        print("Page3")
        self.anim = QPropertyAnimation(self.stackedWidget, b"pos")
        effect = QGraphicsOpacityEffect(self.stackedWidget)
        self.stackedWidget.setGraphicsEffect(effect)
        self.anim.setStartValue(QPoint(200, 100))
        self.anim.setEndValue(QPoint(200, 150))
        self.anim.setEasingCurve(QEasingCurve.OutBounce)
        self.anim.setDuration(500)
        self.anim_2 = QPropertyAnimation(effect, b"opacity")
        self.anim_2.setStartValue(0)
        self.anim_2.setEndValue(1)
        self.anim_2.setDuration(500)
        self.anim_group = QParallelAnimationGroup()
        self.anim_group.addAnimation(self.anim)
        self.anim_group.addAnimation(self.anim_2)
        self.anim_group.start()
        self.stackedWidget.setCurrentWidget(self.page_3)

    def showPage4(self):
        print("Page4")
        self.anim = QPropertyAnimation(self.stackedWidget, b"pos")
        effect = QGraphicsOpacityEffect(self.stackedWidget)
        self.stackedWidget.setGraphicsEffect(effect)
        self.anim.setStartValue(QPoint(200, 100))
        self.anim.setEndValue(QPoint(200, 150))
        self.anim.setEasingCurve(QEasingCurve.OutBounce)
        self.anim.setDuration(500)
        self.anim_2 = QPropertyAnimation(effect, b"opacity")
        self.anim_2.setStartValue(0)
        self.anim_2.setEndValue(1)
        self.anim_2.setDuration(500)
        self.anim_group = QParallelAnimationGroup()
        self.anim_group.addAnimation(self.anim)
        self.anim_group.addAnimation(self.anim_2)
        self.anim_group.start()
        self.stackedWidget.setCurrentWidget(self.page_4)

    def showPage5(self):
        print("Page5")
        self.anim = QPropertyAnimation(self.stackedWidget, b"pos")
        effect = QGraphicsOpacityEffect(self.stackedWidget)
        self.stackedWidget.setGraphicsEffect(effect)
        self.anim.setStartValue(QPoint(200, 100))
        self.anim.setEndValue(QPoint(200, 150))
        self.anim.setEasingCurve(QEasingCurve.OutBounce)
        self.anim.setDuration(500)
        self.anim_2 = QPropertyAnimation(effect, b"opacity")
        self.anim_2.setStartValue(0)
        self.anim_2.setEndValue(1)
        self.anim_2.setDuration(500)
        self.anim_group = QParallelAnimationGroup()
        self.anim_group.addAnimation(self.anim)
        self.anim_group.addAnimation(self.anim_2)
        self.anim_group.start()
        self.stackedWidget.setCurrentWidget(self.page_5)
        
        

       
        


if __name__ == "__main__":
    app = QApplication(sys.argv)
    demo = AppDemo()
    demo.show()
    

    try:
        sys.exit(app.exec_())
    except SystemExit:
        print("closing window...")




