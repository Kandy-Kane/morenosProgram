from typing import List
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.cell import cell
from openpyxl.cell.cell import Cell, ERROR_CODES
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, alignment
from openpyxl.styles.fills import fills
from openpyxl.worksheet.dimensions import SheetDimension
from openpyxl.utils import get_column_letter
import time
import os.path
from os import error, path
# from tkinter import messagebox
import traceback
import sys
import datetime
from PyQt5 import QtCore, QtGui,QtWidgets,uic
from PyQt5.QtWidgets import QMainWindow,QApplication, QWidget
from PyQt5.QtCore import QObject, QThread, pyqtSignal,pyqtSlot
import logging
import threading
import time
import globalVars


thick_border_blue = Border(left=Side(style='thick',color='0066CC'), 
                    right=Side(style='thick',color='0066CC')   
                ) 

thick_border_blue_topBottom = Border(top=Side(style='thick',color='0066CC'), 
                    bottom=Side(style='thick',color='0066CC')   
                )                    

thin_border_all = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin'))

thin_border_all_grey = Border(left=Side(style='thin',color="DDDDDD"), 
                    right=Side(style='thin',color="DDDDDD"), 
                    top=Side(style='thin',color="DDDDDD"), 
                    bottom=Side(style='thin',color="DDDDDD"))

thin_border_sides = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style=None), 
                    bottom=Side(style=None))


thin_border_sides_Bottom = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style=None), 
                    bottom=Side(style='thin'))
ft1 = Font(name='Arial',bold=True, size=14)

align = Alignment(horizontal='center')
thick_border = Border(left=Side(style='thick'), 
                    right=Side(style='thick'), 
                    top=Side(style='thick'), 
                    bottom=Side(style='thick'))
# import main


    #GLOBAL ERROR VARIABLES
fileerrorLabel = None
qualerrorLabel = None
blockerrorLabel = None
starterrorLabel = None
enderrorLabel = None
monthCheck = False


class Worker3(QObject):
    finished = pyqtSignal()
    progress = pyqtSignal(int)
    found = pyqtSignal()


    def ssnSearch(self):
        try:
            print("==========================================STARTING Search=====================================================")


            # global workbook_Title
            workbook_Title = globalVars.page2File
            print(workbook_Title)
            # global workbook
            workbook = load_workbook(filename=workbook_Title)
            # global sheet
            sheet = workbook.active

            rowIndex=2
            columnIndex=1
            lastNameColumnIndex = columnIndex+1
            ssnColumnIndex = columnIndex+2
            dodColumnIndex = columnIndex+3
            flightColumnIndex = columnIndex+4
            signInColumnIndex = columnIndex+5
            signOutColumnIndex = columnIndex+6
            found=False
            firstNamecellref = sheet.cell(row=rowIndex,column=columnIndex)
            lastNamecellref = sheet.cell(row=rowIndex,column=lastNameColumnIndex)
            ssncellref = sheet.cell(row=rowIndex,column=ssnColumnIndex)
            dodCellref = sheet.cell(row=rowIndex,column=dodColumnIndex)
            flightCellRef = sheet.cell(row=rowIndex,column=flightColumnIndex)
            signInCellRef = sheet.cell(row=rowIndex,column=signInColumnIndex)
            signOutCellRef = sheet.cell(row=rowIndex,column=signOutColumnIndex)


            while ssncellref.value:
                if str(ssncellref.value) == str(globalVars.ssn):
                    # firstNamecellref = sheet.cell(rowIndex,column=columnIndex+1)
                    globalVars.returnedFirstName = str(firstNamecellref.value)
                    globalVars.returnedLastName = str(lastNamecellref.value)
                    globalVars.returnedSsn = str(ssncellref.value)
                    print("SSN: "+str(globalVars.returnedSsn))
                    globalVars.returnedDod = str(dodCellref.value)
                    globalVars.returnedFlight = str(flightCellRef.value)
                    globalVars.returnedSignIn = str(signInCellRef.value)
                    globalVars.returnedSignOut = str(signOutCellRef.value)
                    self.found.emit()
                    found =True

                    # time.sleep(1)
                    rowIndex+=1
                else:
                    rowIndex+=1
                    columnIndex=1
                firstNamecellref = sheet.cell(row=rowIndex,column=columnIndex)
                lastNamecellref = sheet.cell(row=rowIndex,column=lastNameColumnIndex)
                ssncellref = sheet.cell(row=rowIndex,column=ssnColumnIndex)
                dodCellref = sheet.cell(row=rowIndex,column=dodColumnIndex)
                flightCellRef = sheet.cell(row=rowIndex,column=flightColumnIndex)
                signInCellRef = sheet.cell(row=rowIndex,column=signInColumnIndex)
                signOutCellRef = sheet.cell(row=rowIndex,column=signOutColumnIndex)
                print(firstNamecellref.coordinate)
            self.finished.emit()

  
        except():
            print(traceback.format_exc())
            # messagebox.showwarning(title="Error Occured", message="something went wrong in EXISTING BOOK. Check your entries and try again")
            if path.exists("errors.txt") == True:
                ct = datetime.datetime.now() 
                with open("errors.txt", "a") as file:
                    file.write("\n"+str(ct)+"\n"+str(traceback.format_exc()))
            else:
                ct = datetime.datetime.now() 
                with open("errors.txt", "x") as file:
                    file.write("\n"+str(ct)+"\n"+str(traceback.format_exc()))







    def dodSearch(self):
        try:
            print("==========================================STARTING Search=====================================================")


            # global workbook_Title
            workbook_Title = globalVars.page2File
            print(workbook_Title)
            # global workbook
            workbook = load_workbook(filename=workbook_Title)
            # global sheet
            sheet = workbook.active

            rowIndex=2
            columnIndex=1
            lastNameColumnIndex = columnIndex+1
            ssnColumnIndex = columnIndex+2
            dodColumnIndex = columnIndex+3
            flightColumnIndex = columnIndex+4
            signInColumnIndex = columnIndex+5
            signOutColumnIndex = columnIndex+6
            found=False
            firstNamecellref = sheet.cell(row=rowIndex,column=columnIndex)
            lastNamecellref = sheet.cell(row=rowIndex,column=lastNameColumnIndex)
            ssncellref = sheet.cell(row=rowIndex,column=ssnColumnIndex)
            dodCellref = sheet.cell(row=rowIndex,column=dodColumnIndex)
            flightCellRef = sheet.cell(row=rowIndex,column=flightColumnIndex)
            signInCellRef = sheet.cell(row=rowIndex,column=signInColumnIndex)
            signOutCellRef = sheet.cell(row=rowIndex,column=signOutColumnIndex)


            while dodCellref.value:
                if str(dodCellref.value) == str(globalVars.dod):
                    # firstNamecellref = sheet.cell(rowIndex,column=columnIndex+1)
                    globalVars.returnedFirstName = str(firstNamecellref.value)
                    globalVars.returnedLastName = str(lastNamecellref.value)
                    globalVars.returnedSsn = str(ssncellref.value)
                    print("SSN: "+str(globalVars.returnedSsn))
                    globalVars.returnedDod = str(dodCellref.value)
                    globalVars.returnedFlight = str(flightCellRef.value)
                    globalVars.returnedSignIn = str(signInCellRef.value)
                    globalVars.returnedSignOut = str(signOutCellRef.value)
                    self.found.emit()
                    found =True

                    # time.sleep(1)
                    rowIndex+=1
                else:
                    rowIndex+=1
                    columnIndex=1
                firstNamecellref = sheet.cell(row=rowIndex,column=columnIndex)
                lastNamecellref = sheet.cell(row=rowIndex,column=lastNameColumnIndex)
                ssncellref = sheet.cell(row=rowIndex,column=ssnColumnIndex)
                dodCellref = sheet.cell(row=rowIndex,column=dodColumnIndex)
                flightCellRef = sheet.cell(row=rowIndex,column=flightColumnIndex)
                signInCellRef = sheet.cell(row=rowIndex,column=signInColumnIndex)
                signOutCellRef = sheet.cell(row=rowIndex,column=signOutColumnIndex)
                print(firstNamecellref.coordinate)
            self.finished.emit()

  
        except():
            print(traceback.format_exc())
            # messagebox.showwarning(title="Error Occured", message="something went wrong in EXISTING BOOK. Check your entries and try again")
            if path.exists("errors.txt") == True:
                ct = datetime.datetime.now() 
                with open("errors.txt", "a") as file:
                    file.write("\n"+str(ct)+"\n"+str(traceback.format_exc()))
            else:
                ct = datetime.datetime.now() 
                with open("errors.txt", "x") as file:
                    file.write("\n"+str(ct)+"\n"+str(traceback.format_exc()))




    def flightSearch(self):
        try:
            print("==========================================STARTING Search=====================================================")


            # global workbook_Title
            workbook_Title = globalVars.page2File
            print(workbook_Title)
            # global workbook
            workbook = load_workbook(filename=workbook_Title)
            # global sheet
            sheet = workbook.active

            rowIndex=2
            columnIndex=1
            lastNameColumnIndex = columnIndex+1
            ssnColumnIndex = columnIndex+2
            dodColumnIndex = columnIndex+3
            flightColumnIndex = columnIndex+4
            signInColumnIndex = columnIndex+5
            signOutColumnIndex = columnIndex+6
            found=False
            firstNamecellref = sheet.cell(row=rowIndex,column=columnIndex)
            lastNamecellref = sheet.cell(row=rowIndex,column=lastNameColumnIndex)
            ssncellref = sheet.cell(row=rowIndex,column=ssnColumnIndex)
            dodCellref = sheet.cell(row=rowIndex,column=dodColumnIndex)
            flightCellRef = sheet.cell(row=rowIndex,column=flightColumnIndex)
            signInCellRef = sheet.cell(row=rowIndex,column=signInColumnIndex)
            signOutCellRef = sheet.cell(row=rowIndex,column=signOutColumnIndex)


            while flightCellRef.value:
                if str(flightCellRef.value).upper() == str(globalVars.flight).upper():
                    # firstNamecellref = sheet.cell(rowIndex,column=columnIndex+1)
                    globalVars.returnedFirstName = str(firstNamecellref.value)
                    globalVars.returnedLastName = str(lastNamecellref.value)
                    globalVars.returnedSsn = str(ssncellref.value)
                    print("SSN: "+str(globalVars.returnedSsn))
                    globalVars.returnedDod = str(dodCellref.value)
                    globalVars.returnedFlight = str(flightCellRef.value)
                    globalVars.returnedSignIn = str(signInCellRef.value)
                    globalVars.returnedSignOut = str(signOutCellRef.value)
                    self.found.emit()
                    found =True

                    # time.sleep(1)
                    rowIndex+=1
                else:
                    rowIndex+=1
                    columnIndex=1
                firstNamecellref = sheet.cell(row=rowIndex,column=columnIndex)
                lastNamecellref = sheet.cell(row=rowIndex,column=lastNameColumnIndex)
                ssncellref = sheet.cell(row=rowIndex,column=ssnColumnIndex)
                dodCellref = sheet.cell(row=rowIndex,column=dodColumnIndex)
                flightCellRef = sheet.cell(row=rowIndex,column=flightColumnIndex)
                signInCellRef = sheet.cell(row=rowIndex,column=signInColumnIndex)
                signOutCellRef = sheet.cell(row=rowIndex,column=signOutColumnIndex)
                print(firstNamecellref.coordinate)
            self.finished.emit()

  
        except():
            print(traceback.format_exc())
            # messagebox.showwarning(title="Error Occured", message="something went wrong in EXISTING BOOK. Check your entries and try again")
            if path.exists("errors.txt") == True:
                ct = datetime.datetime.now() 
                with open("errors.txt", "a") as file:
                    file.write("\n"+str(ct)+"\n"+str(traceback.format_exc()))
            else:
                ct = datetime.datetime.now() 
                with open("errors.txt", "x") as file:
                    file.write("\n"+str(ct)+"\n"+str(traceback.format_exc()))







    def search(self):
        try:
            print("==========================================STARTING Search=====================================================")


            # global workbook_Title
            workbook_Title = globalVars.page2File
            print(workbook_Title)
            # global workbook
            workbook = load_workbook(filename=workbook_Title)
            # global sheet
            sheet = workbook.active

            rowIndex=2
            columnIndex=1
            lastNameColumnIndex = columnIndex+1
            ssnColumnIndex = columnIndex+2
            dodColumnIndex = columnIndex+3
            flightColumnIndex = columnIndex+4
            signInColumnIndex = columnIndex+5
            signOutColumnIndex = columnIndex+6
            found=False
            firstNamecellref = sheet.cell(row=rowIndex,column=columnIndex)
            lastNamecellref = sheet.cell(row=rowIndex,column=lastNameColumnIndex)
            ssncellref = sheet.cell(row=rowIndex,column=ssnColumnIndex)
            dodCellref = sheet.cell(row=rowIndex,column=dodColumnIndex)
            flightCellRef = sheet.cell(row=rowIndex,column=flightColumnIndex)
            signInCellRef = sheet.cell(row=rowIndex,column=signInColumnIndex)
            signOutCellRef = sheet.cell(row=rowIndex,column=signOutColumnIndex)


            while firstNamecellref.value:
                if str(firstNamecellref.value).upper() == str(globalVars.firstName).upper():
                    # firstNamecellref = sheet.cell(rowIndex,column=columnIndex+1)
                    if str(lastNamecellref.value).upper() == str(globalVars.lastName).upper():
                        globalVars.returnedFirstName = str(firstNamecellref.value)
                        globalVars.returnedLastName = str(lastNamecellref.value)
                        globalVars.returnedSsn = str(ssncellref.value)
                        print("SSN: "+str(globalVars.returnedSsn))
                        globalVars.returnedDod = str(dodCellref.value)
                        globalVars.returnedFlight = str(flightCellRef.value)
                        globalVars.returnedSignIn = str(signInCellRef.value)
                        globalVars.returnedSignOut = str(signOutCellRef.value)
                        self.found.emit()
                        found =True

                        # time.sleep(1)
                        rowIndex+=1
                else:
                    rowIndex+=1
                    columnIndex=1
                firstNamecellref = sheet.cell(row=rowIndex,column=columnIndex)
                lastNamecellref = sheet.cell(row=rowIndex,column=lastNameColumnIndex)
                ssncellref = sheet.cell(row=rowIndex,column=ssnColumnIndex)
                dodCellref = sheet.cell(row=rowIndex,column=dodColumnIndex)
                flightCellRef = sheet.cell(row=rowIndex,column=flightColumnIndex)
                signInCellRef = sheet.cell(row=rowIndex,column=signInColumnIndex)
                signOutCellRef = sheet.cell(row=rowIndex,column=signOutColumnIndex)
                print(firstNamecellref.coordinate)
            self.finished.emit()

  
        except():
            print(traceback.format_exc())
            # messagebox.showwarning(title="Error Occured", message="something went wrong in EXISTING BOOK. Check your entries and try again")
            if path.exists("errors.txt") == True:
                ct = datetime.datetime.now() 
                with open("errors.txt", "a") as file:
                    file.write("\n"+str(ct)+"\n"+str(traceback.format_exc()))
            else:
                ct = datetime.datetime.now() 
                with open("errors.txt", "x") as file:
                    file.write("\n"+str(ct)+"\n"+str(traceback.format_exc()))
   

    