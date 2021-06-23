from typing import List
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.cell import cell
from openpyxl.cell.cell import Cell, ERROR_CODES
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, alignment
from openpyxl.styles.fills import fills
from openpyxl.worksheet.dimensions import SheetDimension
from openpyxl.utils import get_column_letter
import time
import os.path
from os import error, path
# from tkinter import messagebox
import traceback
import sys
import datetime
from PyQt5 import QtCore, QtGui,QtWidgets,uic
from PyQt5.QtWidgets import QMainWindow,QApplication, QWidget
from PyQt5.QtCore import QObject, QThread, pyqtSignal,pyqtSlot
import logging
import threading
import time
import globalVars


thick_border_blue = Border(left=Side(style='thick',color='0066CC'), 
                    right=Side(style='thick',color='0066CC')   
                ) 

thick_border_blue_topBottom = Border(top=Side(style='thick',color='0066CC'), 
                    bottom=Side(style='thick',color='0066CC')   
                )                    

thin_border_all = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin'))

thin_border_all_grey = Border(left=Side(style='thin',color="DDDDDD"), 
                    right=Side(style='thin',color="DDDDDD"), 
                    top=Side(style='thin',color="DDDDDD"), 
                    bottom=Side(style='thin',color="DDDDDD"))

thin_border_sides = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style=None), 
                    bottom=Side(style=None))


thin_border_sides_Bottom = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style=None), 
                    bottom=Side(style='thin'))
ft1 = Font(name='Arial',bold=True, size=14)

align = Alignment(horizontal='center')
thick_border = Border(left=Side(style='thick'), 
                    right=Side(style='thick'), 
                    top=Side(style='thick'), 
                    bottom=Side(style='thick'))
# import main


    #GLOBAL ERROR VARIABLES
fileerrorLabel = None
qualerrorLabel = None
blockerrorLabel = None
starterrorLabel = None
enderrorLabel = None
monthCheck = False


class Worker2(QObject):
    finished = pyqtSignal()
    progress = pyqtSignal(int)

    def addEntry(self):
        try:
            print("==========================================STARTING ADD ENTRY=====================================================")


            # global workbook_Title
            workbook_Title = globalVars.page1File
            print(workbook_Title)
            # global workbook
            workbook = load_workbook(filename=workbook_Title)
            # global sheet
            sheet = workbook.active

            rowIndex=2
            columnIndex=1
            cellref = sheet.cell(row=rowIndex,column=columnIndex)

            while cellref.value:
                rowIndex+=1
                cellref = sheet.cell(row=rowIndex,column=columnIndex)
                print(cellref.coordinate)

            cellref.value = globalVars.firstName
            columnIndex+=1
            cellref = sheet.cell(row=rowIndex,column=columnIndex)

            cellref.value = globalVars.lastName
            columnIndex+=1
            cellref = sheet.cell(row=rowIndex,column=columnIndex)

            cellref.value = globalVars.ssn
            columnIndex+=1
            cellref = sheet.cell(row=rowIndex,column=columnIndex)

            cellref.value = globalVars.dod
            columnIndex+=1
            cellref = sheet.cell(row=rowIndex,column=columnIndex)

            cellref.value = globalVars.flight
            columnIndex+=1
            cellref = sheet.cell(row=rowIndex,column=columnIndex)

            cellref.value = globalVars.signInDate
            columnIndex+=1
            cellref = sheet.cell(row=rowIndex,column=columnIndex)

            cellref.value = globalVars.signOutDate
            columnIndex+=1
            cellref = sheet.cell(row=rowIndex,column=columnIndex)
            workbook.save(filename=workbook_Title)
            self.finished.emit()



           
        except():
            print(traceback.format_exc())
            # messagebox.showwarning(title="Error Occured", message="something went wrong in EXISTING BOOK. Check your entries and try again")
            # if path.exists("errors.txt") == TRUE:
            #     ct = datetime.datetime.now() 
            #     with open("errors.txt", "a") as file:
            #         file.write("\n"+str(ct)+"\n"+str(traceback.format_exc()))
            # else:
            #     ct = datetime.datetime.now() 
            #     with open("errors.txt", "x") as file:
            #         file.write("\n"+str(ct)+"\n"+str(traceback.format_exc()))
   

    